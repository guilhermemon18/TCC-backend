from openpyxl import Workbook

from openpyxl import load_workbook


def cria_coluna(arquivo_excel, planilha, name_other_planilha):
    max_linha = planilha.max_row
    coluna_notac1 = planilha.max_column + 1
    disciplina = arquivo_excel[name_other_planilha]
    planilha.cell(row=1, column=coluna_notac1).value = name_other_planilha
    for i in range(1 + 1, disciplina.max_row + 1):
        for k in range(1 + 1, max_linha + 1):
            if disciplina.cell(row=i, column=1).value == planilha.cell(row=k, column=1).value:
                planilha.cell(row=k, column=coluna_notac1).value = disciplina.cell(row=i, column=2).value


def main():
    # Carrega um arquivo do excel.
    caminho = 'guilherme 0627.xlsx'
    arquivo_excel = load_workbook(caminho)

    dadosFinais = arquivo_excel['dadosFinais']
    copia = arquivo_excel.copy_worksheet(dadosFinais)
    notacomp1 = arquivo_excel['nota comp1']

    max_linha = dadosFinais.max_row
    max_coluna = dadosFinais.max_column

    # cria a coluna de notas de comp1.
    coluna_notac1 = dadosFinais.max_column + 1

    dadosFinais.cell(row=1, column=coluna_notac1).value = 'notaComp1'
    for i in range(1 + 1, notacomp1.max_row + 1):
        for k in range(1 + 1, max_linha + 1):
            if notacomp1.cell(row=i, column=1).value == dadosFinais.cell(row=k, column=1).value:
                dadosFinais.cell(row=k, column=coluna_notac1).value = notacomp1.cell(row=i, column=2).value

    cria_coluna(arquivo_excel, dadosFinais, 'nota calc')
    cria_coluna(arquivo_excel, dadosFinais, 'nota prat desp')
    cria_coluna(arquivo_excel, dadosFinais, 'nota fisica')
    cria_coluna(arquivo_excel, dadosFinais, 'nota int adm')
    cria_coluna(arquivo_excel, dadosFinais, 'nota GA')
    cria_coluna(arquivo_excel, dadosFinais, 'nota ingles')
    cria_coluna(arquivo_excel, dadosFinais, 'nota PE')
    cria_coluna(arquivo_excel, dadosFinais, 'nota LMD')
    cria_coluna(arquivo_excel, dadosFinais, 'nota ICC')
    cria_coluna(arquivo_excel, dadosFinais, 'nota met c')
    cria_coluna(arquivo_excel, dadosFinais, 'nota sociologia')
    cria_coluna(arquivo_excel, dadosFinais, 'nota tec redacao')
    cria_coluna(arquivo_excel, dadosFinais, 'desv pad')


    max_linha = dadosFinais.max_row
    max_coluna = dadosFinais.max_column
    for i in range(1, max_linha + 1):
        for j in range(1, max_coluna + 1):
            print(dadosFinais.cell(row=i, column=j).value, end=" - ")
        print('\n')

    arquivo_excel.save('planilha.xlsx')

    # instancia um novo arquivo excel
    arquivo_excel = Workbook()

    # obtem a planilha ativa padrao que é criada junto com o arquivo excel "Sheet1"
    planilha1 = arquivo_excel.active

    # renomeia para um nome de nossa preferência
    planilha1.title = "planilha_dados_juntos"

    # salva o arquivo do excel criado
    arquivo_excel.save("transformado.xlsx")

main()
#wb = load_workbook('dados final guilherme_SEM RA.xlsx')
#for sheet in wb:
 #   print(sheet)

#ws = wb['DADOS TREINO SEM OS NÃO REGS']
#print(ws['A2'].value)