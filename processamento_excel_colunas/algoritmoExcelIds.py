# Carrega um arquivo do excel.
from openpyxl.reader.excel import load_workbook
from pathlib import Path, PureWindowsPath
from tkinter import filedialog, messagebox
import tkinter as tk
import os

caminho = Path.home() / "Desktop"
caminho = str(caminho)
print("Caminho principal da área de trabalho:" + caminho)

root = tk.Tk()
root.withdraw()

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
file_types = [("arquivos de Excel", "*.xlsx")]
file_path_gr73 = filedialog.askopenfilename(initialdir=caminho, filetypes=file_types,
                                            title="Selecione o arquivo GR 73_geracao id.xlsx"
                                            , initialfile="GR 73_geracao id.xlsx")
if file_path_gr73:
    print(file_path_gr73)

file_path_gr30 = filedialog.askopenfilename(initialdir=caminho, filetypes=file_types,
                                            title="Selecione o arquivo GR 30_geração ID.xlsx"
                                            , initialfile="GR 30_geração ID.xlsx")
if file_path_gr30:
    print(file_path_gr30)

excel_gr73 = load_workbook(file_path_gr73)
excel_gr30 = load_workbook(file_path_gr30)

planilha1gr73 = excel_gr73['Planilha1']
planilha1gr30 = excel_gr30['Planilha1']

max_linha = planilha1gr73.max_row


for i in range(1 + 1, planilha1gr73.max_row + 1):
    for k in range(1 + 1, planilha1gr30.max_row + 1):

        if int(planilha1gr73.cell(row=i, column=16).value) == int(planilha1gr30.cell(row=k, column=2).value):
            #print(planilha1gr73.cell(row=i, column=16).value)
            #print(planilha1gr30.cell(row=k, column=2).value)
            #print("Deu Igual!")
            planilha1gr30.cell(row=k, column=1).value = planilha1gr73.cell(row=i, column=17).value
            #break


excel_gr30.save(file_path_gr30)


messagebox.showwarning("Finalizado", "O arquivo GR73 está pronto com o mesmo nome no mesmo local!")
